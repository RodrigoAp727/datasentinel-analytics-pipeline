from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from analysis.anomaly_detector import AnomalyFinding, AnomalyReport
from reporting.report_generator import ReportGenerator
from validation.quality_checker import QualityIssue, QualityReport

TIMESTAMP_PATTERN = re.compile(r"^report_\d{8}_\d{6}\.(xlsx|html)$")


def _sample_quality_report() -> QualityReport:
    return QualityReport(
        status="WARNING",
        score=82.5,
        issues=[
            QualityIssue(
                check="null_threshold",
                severity="WARNING",
                message="Null ratio above threshold",
                affected_columns=["value"],
                details={"ratio": 0.3, "threshold": 0.2},
            )
        ],
    )


def _sample_anomaly_report() -> AnomalyReport:
    return AnomalyReport(
        anomalies=[
            AnomalyFinding(
                column="value",
                index=3,
                severity="alta",
                method="zscore",
                value=999.0,
                score=4.5,
                message="Outlier detected",
            )
        ]
    )


def test_generate_excel_report_creates_expected_sheets(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "id": [1, 2, 3],
            "value": [10.0, 20.0, 30.0],
            "category": ["A", "B", "C"],
        }
    )

    generator = ReportGenerator(
        dataframe=df,
        quality_report=_sample_quality_report(),
        anomaly_report=_sample_anomaly_report(),
        reports_dir=tmp_path / "reports",
    )

    excel_path = generator.generate_excel_report()

    assert excel_path.exists()
    assert TIMESTAMP_PATTERN.match(excel_path.name)

    workbook = load_workbook(excel_path)
    expected_sheets = {
        "Resumo Executivo",
        "Perfil da Base",
        "Qualidade de Dados",
        "Anomalias Detectadas",
        "Dados Brutos",
    }
    assert expected_sheets.issubset(set(workbook.sheetnames))


def test_generate_html_report_creates_file_with_plotly_content(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "timestamp": pd.date_range("2026-01-01", periods=5, freq="D"),
            "sales": [100.0, 101.0, 99.0, 150.0, 102.0],
            "machine_load": [60.0, 62.0, 59.0, 61.0, 58.0],
        }
    )

    generator = ReportGenerator(
        dataframe=df,
        quality_report=_sample_quality_report(),
        anomaly_report=_sample_anomaly_report(),
        reports_dir=tmp_path / "reports",
    )

    html_path = generator.generate_html_report()

    assert html_path.exists()
    assert TIMESTAMP_PATTERN.match(html_path.name)

    content = html_path.read_text(encoding="utf-8")
    assert "plotly" in content.lower()
    assert "Resumo Executivo" in content
    assert "Relatorio Executivo" in content
    assert "Relatorio Escrito da Planilha" in content
    assert "Perfil Detalhado da Base" in content
    assert "Problemas de Qualidade" in content
    assert "Anomalias Detectadas" in content
    assert "Distribuicao de Anomalias por Severidade" in content


def test_generate_reports_creates_both_artifacts_and_reports_directory(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "id": [1, 2, 3, 4],
            "value": [10.0, None, 30.0, 35.0],
        }
    )

    reports_dir = tmp_path / "reports"
    generator = ReportGenerator(
        dataframe=df,
        quality_report=_sample_quality_report(),
        anomaly_report=_sample_anomaly_report(),
        reports_dir=reports_dir,
    )

    paths = generator.generate_reports()

    assert reports_dir.exists()
    assert paths["excel"].exists()
    assert paths["html"].exists()
    assert TIMESTAMP_PATTERN.match(paths["excel"].name)
    assert TIMESTAMP_PATTERN.match(paths["html"].name)
